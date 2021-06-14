import os
import sys
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook, workbook, load_workbook
from openpyxl.utils import get_column_letter


time = []
series = []

filepath = sys.argv[1]

def readxlxx(filePath):

    fp = os.getcwd()

    fp = fp + f"/{filepath}"
    
    # Checking file exists
    if not os.path.isfile(fp):
        print("Error - specified folderpath is not a directory. Use option d to see available folderpaths.")

    try:
        workbook = load_workbook(fp)
        sheet = workbook.active
        
        currRow = 1
        currTime = 0
        for row in sheet.values:
            if currRow == 7:
                for val in row:
                    series.append(val)
                    time.append(currTime)
                    currTime += 1
            currRow += 1
        
        # print(series)
        # print(time)


    except Exception as e:
        print(e)
        exit(-1)



def plot_series(time, series, format="-", start=0, end=None):
    plt.plot(time[start:end], series[start:end], format)
    plt.xlabel("Time")
    plt.ylabel("Value")
    plt.grid(True)

readxlxx(filepath)
plt.figure(figsize=(10, 6))
plot_series(np.array(time), np.array(series))
plt.show()
